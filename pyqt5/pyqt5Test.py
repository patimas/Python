import sys, os, pickle   # 파일 입출력 처리
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QAction, QDesktopWidget, QGridLayout, \
                            QLabel, QLineEdit, QVBoxLayout, QTabWidget, QMenu, QPushButton, QHBoxLayout, QGroupBox, \
                            QMessageBox, QTableWidget, QTableWidgetItem, QCheckBox, QComboBox, QFileDialog, \
                            QSlider, QRadioButton, QTextEdit
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import QDate, QTime, QTimer, Qt
from datetime import timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

class MyApp(QMainWindow):

    ############################################################
    # 초기화
    ############################################################
    def __init__(self, parent=None):
        super().__init__(parent)
        self._configSetting()
        self.initUI()


    ############################################################
    # Setting
    ############################################################
    def _configSetting(self):
        self.imgPath = 'pyqt5\img\\'

    ############################################################
    # UI구성
    ############################################################
    def initUI(self):
        
        self.setWindowTitle('PyQt5 GUI')
        self.setGeometry(200, 200, 450,350)
        # self.setWindowIcon(QIcon(imgPath+'tray.ico'))  # Application Icon Setting
        
        #  현재 날짜 설정
        self.date = QDate.currentDate()
        self.statusBar().showMessage(self.date.toString(Qt.DefaultLocaleLongDate))

        # self.centralWidget = QLabel('Hello Workd!')
        # self.centralWidget.setAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        # self.setCentralWidget(self.centralWidget)

        self._Actions()
        self._MenuBar()
        self._ToolBar()
        self._ConnectActions()
        # self._ContextMenu()

        self._OpneConfig()

        # self._CenterArea()   # 창 가운데 정렬

        # Tab 호출
        self.TabWidget = TabWidget(self)
        self.setCentralWidget(self.TabWidget)

        self._LoadSetting()
        # self.tabWidget = TabWidget(self)


    ############################################################
    #
    #  설정 파일 오픈 (config.ini)
    #
    ############################################################
    def _OpneConfig(self):
        configPath = "./pyqt5/conf"
        configFile = "config.ini"
        self.configFileFullPath = configPath+"/"+configFile

        if os.path.isdir(configPath) == False:
            os.mkdir("pyqt5/conf")


        if os.path.isfile(self.configFileFullPath) == True:
            f = open(self.configFileFullPath,"rb")
            self.configValueRead = pickle.load(f)
            f.close()
        else:
            self.configValueRead = ""

    ############################################################
    # Action 생성
    ############################################################
    def _Actions(self):
        self.openAction = QAction(QIcon(self.imgPath+'open.png'), '&Open',self)
        self.saveAction = QAction(QIcon(self.imgPath+'save.png'), '&Save',self)

        self.exitAction = QAction(QIcon(self.imgPath+'exit.png'), '&Exit',self)
        self.settingAction = QAction(QIcon(self.imgPath+'setting.png'), '&Setting',self)

        self.openAction.setShortcut('Ctrl+O')
        self.saveAction.setShortcut('Ctrl+S')

    ############################################################
    # Menu 생성
    ############################################################
    def _MenuBar(self):

        menuBar = self.menuBar()

        fileMenu = menuBar.addMenu('&File')
        editMenu = menuBar.addMenu("&Edit")
        helpMenu = menuBar.addMenu("&Help")

        fileMenu.addAction(self.openAction)
        fileMenu.addAction(self.saveAction)
        fileMenu.addAction(self.exitAction)



        helpMenu.addAction(self.settingAction)

    def _ToolBar(self):
        fileToolBar = self.addToolBar('File')
        fileToolBar.addAction(self.openAction)
        fileToolBar.addAction(self.saveAction)
        fileToolBar.addAction(self.exitAction)
        fileToolBar.addAction(self.settingAction)

        editToolBar = self.addToolBar('Edit')


    # def _ContextMenu(self):

    #     # 위에서 생성된 Label에 적용 한다.
    #     self.centralWidget.setContextMenuPolicy(Qt.ActionsContextMenu)

    #     self.centralWidget.addAction(self.openAction)
    #     self.centralWidget.addAction(self.saveAction)
    #     self.centralWidget.addAction(self.exitAction)
    #     self.centralWidget.addAction(self.settingAction)


    ############################################################
    # Context 생성
    ############################################################
    # Qt는 오른쪽 마우스 클릭이 일어나면 mousePressEvent(), mouseReleaseEvent(), contextMenuEvent()의 순서로 이벤트를 발생시키게 된다. 
    # 컨텍스트메뉴를 처리하려면 contextMenuEvent() 핸들러를 오버라이드(override)하여 처리하면 된다.
    def contextMenuEvent(self, event):
        menu = QMenu(self.centralWidget)

        separator = QAction(self)
        separator.setSeparator(True)

        menu.addAction(self.openAction)
        menu.addAction(self.saveAction)
        menu.addAction(self.exitAction)
        menu.addAction(self.settingAction)

        menu.addAction(separator)



        # 사용자가 화면에 클릭을 하면 마우스 포인트를 반환 한다.
        menu.exec(event.globalPos())


    ############################################################
    # 
    ############################################################
    def _ConnectActions(self):
        self.openAction.triggered.connect(self._Open)
        self.saveAction.triggered.connect(self._SaveSetting)
        self.exitAction.triggered.connect(self.close)

    ############################################################
    # Event 함수
    ############################################################
    def _Open(self):
        self.centralWidget.setText('Event Test!')


    # 화면을 항상 중앙으로
    def _CenterArea(self):
        qr = self.frameGeometry()       # 창의 위치와 크기 정보를 가져온다.
        cp = QDesktopWidget().availableGeometry().center()  # 사용하는 모니터 화면의 가운데 위치를 파악한다.
        qr.moveCenter(cp)               # 창의 직사각형 위치를 화면의 중심의 위치로 이동한다.
        self.move(qr.topLeft())         # 현재 창을, 화면의 중심으로 이동했던 직사각형(qr)의 위치로 이동시킨다.


    def _SaveSetting(self):


        print('Config Save Start!!')
        configValueInput = {}

        # tab2
        for i in range(9):
            print(self.TabWidget.tab2LineEdit[i].text())
            configValueInput["tab2LineEdit"+str(i).zfill(2)] = self.TabWidget.tab2LineEdit[i].text()
            


        # tab3
        for i in range(self.TabWidget.tab3Table.rowCount()):
            for j in range(self.TabWidget.tab3TableColumn):
                if j > 0 :
                    print(self.TabWidget.tab3Table.item(i,j).text())
                    configValueInput["tab3Table"+str(i)+str(j)] = self.TabWidget.tab3Table.item(i,j).text()


        # Tab3Table 의 전체 Row수를 별도로 기입 한다.
        configValueInput['tab3TableTotalCount'] = self.TabWidget.tab3Table.rowCount()

        # tab8
        configValueInput['tab8TextEdit01'] = self.TabWidget.tab8TextEdit01.toPlainText()
        configValueInput['tab8TextEdit02'] = self.TabWidget.tab8TextEdit02.toPlainText()

        f = open(self.configFileFullPath,'wb')
        pickle.dump(configValueInput, f)
        f.close()

        QMessageBox.about(self, '저장', '저장 완료!')

        print('Config Save End!!')

    def _LoadSetting(self):
        # self.TabWidget
        print('Config Load Start!!!')

        print(self.configValueRead)
        try:

            # tab2
            for i in range(9):
                if 'tab2LineEdit'+str(i).zfill(2)  in self.configValueRead:
                    self.TabWidget.tab2LineEdit[i].setText( self.configValueRead['tab2LineEdit'+str(i).zfill(2)] )
                    
                    
            # tab3
            if 'tab3TableTotalCount' in self.configValueRead:
                tab3TableMaxCount = self.configValueRead['tab3TableTotalCount']
                self.TabWidget.tab3Table.setRowCount(tab3TableMaxCount)

                for i in range(tab3TableMaxCount):
                    self.TabWidget.tab3Table.setCellWidget(i, 0, QCheckBox())
                    for j in range(self.TabWidget.tab3TableColumn):
                        if j > 0:
                            print('tableWidjet : '+self.configValueRead['tab3Table'+str(i)+str(j)])
                            self.TabWidget.tab3Table.setItem(i, j, QTableWidgetItem(self.configValueRead['tab3Table'+str(i)+str(j)]))

            # tab8
            self.TabWidget.tab8TextEdit01.setText(self.configValueRead['tab8TextEdit01'])
            self.TabWidget.tab8TextEdit02.setText(self.configValueRead['tab8TextEdit02'])

            print('Config Load End!!!')
        except:
            print("Error Bye~~~~")
        finally:
            print("End Bye~~~~")


# Tab 설정
class TabWidget(QWidget):

    def __init__(self, parent):
        super(QWidget, self).__init__(parent)
        self.layout = QVBoxLayout()

        # Tab 스크린 설정
        self.tabs = QTabWidget()
        self.tab1 = QWidget()
        self.tab2 = QWidget()
        self.tab3 = QWidget()
        self.tab8 = QWidget()
        self.tab9 = QWidget()  # 설정

        # Tab 추가
        self.tabs.addTab(self.tab1, '시스템')
        self.tabs.addTab(self.tab2, '엑셀')
        self.tabs.addTab(self.tab3, '쭈~')
        self.tabs.addTab(self.tab8, '반복')
        self.tabs.addTab(self.tab9, '기타')

        # Tab 설정
        self._tab1()   # 시스템 종료
        self._tab2()   # 엑셀(데이터 설계서 만들기 매크로)
        self._tab3()   # 그리드 데이터 표
        self._tab8()   # 프로시저 일괄 생성
        self._tab9()   # 기타 기능

        # 위젯에 탭 추가
        self.layout.addWidget(self.tabs)
        self.setLayout(self.layout)

    def _tab1(self):

        tab1OutLayout = QHBoxLayout()

        tab1LeftLayout = QVBoxLayout()
        tab1RightLayout = QVBoxLayout()


        # 1.시스템 종료 설정
        tab1LeftGroupBox = QGroupBox('시스템 종료')
        self.shutdownBtn = QPushButton('종료')
        self.shutdownCancelBtn = QPushButton('취소')
        self.tab1ComboBox01 = QComboBox()
        self.tab1ComboBox02 = QComboBox()
        self.tab1ComboBox03 = QComboBox()
        self.tab1Label01 = QLabel()
        self.tab1Slider = QSlider(Qt.Horizontal)
        self.tab1Slider.setTickInterval(20)
        self.tab1Slider.setTickPosition(QSlider.TicksAbove)
        self.tab1SliderLabel = QLabel()
        self.tab1RadioBtn01 = QRadioButton()
        self.tab1RadioBtn02 = QRadioButton()



        tab1TimeHarf = ['오전','오후']
        tab1TimeHour = ['00','01','02','03','04','05','06','07','08','09','10','11']
        tab1TimeMinute = ['00','10','20','30','40','50']

        tab1LeftGrid = QGridLayout()
        tab1LeftGrid.addWidget(self.shutdownBtn, 0, 1)
        tab1LeftGrid.addWidget(self.shutdownCancelBtn, 0, 2)

        tab1LeftGrid.addWidget(self.tab1RadioBtn01, 1, 0)
        tab1LeftGrid.addWidget(self.tab1Slider, 1, 1, 1, 2)  # row, column, 행, 열
        tab1LeftGrid.addWidget(self.tab1SliderLabel, 1, 3)  # row, column, 행, 열

        tab1LeftGrid.addWidget(self.tab1RadioBtn02, 2, 0)
        tab1LeftGrid.addWidget(self.tab1ComboBox01, 2, 1)
        tab1LeftGrid.addWidget(self.tab1ComboBox02, 2, 2)
        tab1LeftGrid.addWidget(self.tab1ComboBox03, 2, 3)
        
        tab1LeftGrid.addWidget(self.tab1Label01, 0, 3)  # row, column, 행, 열
        

        self.tab1ComboBox01.addItems(tab1TimeHarf)
        self.tab1ComboBox02.addItems(tab1TimeHour)
        self.tab1ComboBox03.addItems(tab1TimeMinute)

        self.tab1RadioBtn01.setChecked(True)
        self.shutdownBtn.clicked.connect(self._shutdown)
        self.shutdownCancelBtn.clicked.connect(self._shutdownCancel)
        self.tab1Slider.valueChanged.connect(self._tab1SliderLabel)

        # self.tab1ComboBox01.currentIndexChanged.connect(self._tab1ComboBoxEvent)  # ComboBox 선택 이벤트


        tab1OutLayout.addLayout(tab1LeftLayout)
        tab1OutLayout.addLayout(tab1RightLayout)
        tab1LeftLayout.addWidget(tab1LeftGroupBox)
        tab1LeftGroupBox.setLayout(tab1LeftGrid)

        tab1OutLayout.addStretch()
        tab1LeftLayout.addStretch()

        self.tab1.setLayout(tab1OutLayout)




    def _tab2(self):
        # QVBoxLayout, QHBoxLayout, QBoxLayout, QGridLayout, QLayout
        # self.tab2.layout = QVBoxLayout(self)
        # # self.pushButton1 = QPushButton('버튼')
        # self.tab2.layout.addWidget(tab2Grid)
        
        tab2OutLayout = QHBoxLayout()

        tab2LeftLayout = QVBoxLayout()
        tab2RightLayout = QVBoxLayout()


        self.testBtn = QPushButton('버튼')
        tab2LeftLayout.addWidget(self.testBtn)
        self.testBtn.clicked.connect(self._print)


        tab2LeftGroupBox = QGroupBox('파일 경로')

        # LineEdit List 생성
        self.tab2LineEdit = []
        for i in range(9):
            self.tab2LineEdit.append(QLineEdit())

        tab2LeftGrid = QGridLayout()
        tab2LeftGrid.addWidget(QLabel('경   로 : '), 0, 0)
        tab2LeftGrid.addWidget(QLabel('소스파일 : '), 1, 0)
        tab2LeftGrid.addWidget(QLabel('타겟파일 : '), 2, 0)


        tab2LeftGrid.addWidget(self.tab2LineEdit[0], 0, 1)
        tab2LeftGrid.addWidget(self.tab2LineEdit[1], 1, 1)
        tab2LeftGrid.addWidget(self.tab2LineEdit[2], 2, 1)

       
        tab2RightGroupBox = QGroupBox('옵션')

        tab2RightGrid = QGridLayout()
        tab2RightGrid.addWidget(QLabel('시   작 row : '), 0, 0)
        tab2RightGrid.addWidget(QLabel('데이터 범위 : '), 1, 0)
        tab2RightGrid.addWidget(QLabel('샘플Sheet : '), 2, 0)


        tab2RightGrid.addWidget(self.tab2LineEdit[3], 0, 1)
        tab2RightGrid.addWidget(self.tab2LineEdit[4], 1, 1)
        tab2RightGrid.addWidget(self.tab2LineEdit[5], 2, 1)


        tab2RightGrid.addWidget(QLabel('append row : '), 0, 2)
        tab2RightGrid.addWidget(QLabel('소스파일 : '), 1, 2)
        tab2RightGrid.addWidget(QLabel('타겟파일 : '), 2, 2)

        tab2RightGrid.addWidget(self.tab2LineEdit[6], 0, 3)
        tab2RightGrid.addWidget(self.tab2LineEdit[7], 1, 3)
        tab2RightGrid.addWidget(self.tab2LineEdit[8], 2, 3)

        tab2OutLayout.addLayout(tab2LeftLayout)
        tab2OutLayout.addLayout(tab2RightLayout)

        tab2LeftLayout.addWidget(tab2LeftGroupBox)
        tab2LeftGroupBox.setLayout(tab2LeftGrid)

        tab2RightLayout.addWidget(tab2RightGroupBox)
        tab2RightGroupBox.setLayout(tab2RightGrid)


        tab2OutLayout.addStretch()
        tab2LeftLayout.addStretch()
        tab2RightLayout.addStretch()
        
        self.tab2.setLayout(tab2OutLayout)



    def _tab3(self):
        
        tab3OutLayout = QVBoxLayout()

        tab3ToptLayout = QHBoxLayout()
        tab3BottonLayout = QHBoxLayout()

        self.insertBtn = QPushButton('입력')
        self.deleteBtn = QPushButton('삭제')
        tab3ToptLayout.addWidget(self.insertBtn)
        tab3ToptLayout.addWidget(self.deleteBtn)

        self.insertBtn.clicked.connect(self._InsertTableRow)
        self.deleteBtn.clicked.connect(self._DeleteTableRow)

       
        self.tab3TableColumn = 7

        # self.tab3Table = QTableWidget(self.tab3TableRow, self.tab3TableColumn)
        self.tab3Table = QTableWidget()

        self.tab3Table.setColumnCount(self.tab3TableColumn)
        self.tab3Table.setHorizontalHeaderLabels(['','가','나','다','라','마','바'])
        self.tab3Table.resizeColumnsToContents()
        self.tab3Table.resizeRowsToContents()

        # self.tab3Table.item(0,0).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        # self.tab3Table.setFixedSize(500,400)
        # self.tab3Table.setSortingEnabled(True)  # 헤더를 클릭하여 정렬

        # self.tab3Table.setEditTriggers(QAbstractItemView.AllEditTriggers)   # 수정 가능 하게

        tab3OutLayout.addLayout(tab3ToptLayout)
        tab3OutLayout.addLayout(tab3BottonLayout)

        tab3BottonLayout.addWidget(self.tab3Table)
        # tab3RightGroupBox.addWidget(self.tab3Table)


        # tab3OutLayout.addStretch()
        # tab3LeftLayout.addStretch()
        # tab3RightLayout.addStretch()
        
        self.tab3.setLayout(tab3OutLayout)

    def _tab8(self):
        tab8OutLayout = QHBoxLayout()

        tab8LeftLayout = QVBoxLayout()
        tab8RightLayout = QVBoxLayout()

        # Group Box 1
        tab8GroupBox1 = QGroupBox('소스')
        tab8InnerLayout1 = QHBoxLayout()
        self.tab8FilePath = QLineEdit()
        self.tab8FilePathdBtn = QPushButton('찾기')
        self.tab8ActionBtn = QPushButton('시작')
        tab8InnerLayout1.addWidget(self.tab8FilePath)
        tab8InnerLayout1.addWidget(self.tab8FilePathdBtn)
        tab8InnerLayout1.addWidget(self.tab8ActionBtn)

        # Group Box 2
        tab8GroupBox2 = QGroupBox('Text')
        tab8InnerLayout2 = QVBoxLayout()
        self.tab8TextEdit01 = QTextEdit()
        self.tab8TextEdit02 = QTextEdit()
        self.tab8TextEdit01.setAcceptRichText(False)  # 모두 플레인 텍스트로 인식합니다.
        self.tab8TextEdit02.setAcceptRichText(False)  # 모두 플레인 텍스트로 인식합니다.
        tab8InnerLayout2.addWidget(self.tab8TextEdit01)
        tab8InnerLayout2.addWidget(self.tab8TextEdit02)


        # Action Event
        self.tab8FilePathdBtn.clicked.connect(lambda: self._fileOpen(self.tab8FilePath))
        self.tab8ActionBtn.clicked.connect(self._tab8ExcelCreate)


        # Set Layout
        tab8OutLayout.addLayout(tab8LeftLayout)
        tab8OutLayout.addLayout(tab8RightLayout)
        tab8LeftLayout.addWidget(tab8GroupBox1)
        tab8LeftLayout.addWidget(tab8GroupBox2)

        tab8GroupBox1.setLayout(tab8InnerLayout1)
        tab8GroupBox2.setLayout(tab8InnerLayout2)

        self.tab8.setLayout(tab8OutLayout)


    def _tab9(self):
        tab9OutLayout = QHBoxLayout()

        tab9LeftLayout = QVBoxLayout()
        tab9RightLayout = QVBoxLayout()

        # Group Box 1
        tab9LeftGroupBox1 = QGroupBox('파일 경로')
        tab9LeftInnerLayout1 = QHBoxLayout()
        self.tab9FilePath = QLineEdit()
        self.tab9FilePathdBtn = QPushButton('찾기')
        tab9LeftInnerLayout1.addWidget(self.tab9FilePath)
        tab9LeftInnerLayout1.addWidget(self.tab9FilePathdBtn)


        # Group Box 2
        tab9LeftGroupBox2 = QGroupBox('Action')
        tab9LeftInnerLayout2 = QHBoxLayout()
        self.tab9Label1 = QLabel('단어 찾기 : ')
        self.tab9Keyword = QLineEdit()
        self.tab9KeywordSearchBtn = QPushButton('시작')
        tab9LeftInnerLayout2.addWidget(self.tab9Label1)
        tab9LeftInnerLayout2.addWidget(self.tab9Keyword)
        tab9LeftInnerLayout2.addWidget(self.tab9KeywordSearchBtn)


        # Action Event
        self.tab9FilePathdBtn.clicked.connect(lambda: self._fileOpen(self.tab9FilePath))
        self.tab9KeywordSearchBtn.clicked.connect(self._tab9KeywordSearch)


        # Set Layout
        tab9OutLayout.addLayout(tab9LeftLayout)
        tab9OutLayout.addLayout(tab9RightLayout)
        tab9LeftLayout.addWidget(tab9LeftGroupBox1)
        tab9LeftLayout.addWidget(tab9LeftGroupBox2)

        tab9LeftGroupBox1.setLayout(tab9LeftInnerLayout1)
        tab9LeftGroupBox2.setLayout(tab9LeftInnerLayout2)


        tab9OutLayout.addStretch()
        tab9LeftLayout.addStretch()

        self.tab9.setLayout(tab9OutLayout)

    ########################################################################
    #        공용 이벤트 엑션  
    ########################################################################
    def _fileOpen(self, value):
        # 파일을 읽어 특정 문자열이 있는 라인을 추출 한다.
        fileOpen = QFileDialog.getOpenFileName()
        value.setText(fileOpen[0])


    ########################################################################
    #        각 TAB 별 이벤트 엑션  
    ########################################################################
    # Tab1
    def _shutdown(self):
        currtime = QTime.currentTime()
        settime1 = self.tab1ComboBox01.currentText()
        settime2 = self.tab1ComboBox02.currentText()
        settime3 = self.tab1ComboBox03.currentText()

        if self.tab1RadioBtn01.isChecked() == True:
            if self.tab1SliderLabel.text() == "":
                QMessageBox.about(self, '저장', '시간 선택 안됨!')
                return

            self.countTime = int(self.tab1SliderLabel.text()) * 60

        if self.tab1RadioBtn02.isChecked() == True:
            if settime1 == '오후':
                settime2 = str(int(settime2) + 12)

            self.countTime = currtime.secsTo(QTime.fromString(settime2+settime3+'00','hhmmss'))

            if self.countTime < 0 :
                self.countTime = self.countTime + 86400


        self.timer = QTimer()
        self.timer.start(1000)    # 1000 = 1초
        self.timer.timeout.connect(self._startTimer)


    def _shutdownCancel(self):
        self._sotpTimer()

        # os.system('shutdown -a')

    def _startTimer(self):
        self.tab1Label01.setText(str(timedelta(seconds=self.countTime)))
        self.countTime = self.countTime -1

        if self.countTime == 0:
            self._sotpTimer()
            os.system('shutdown -s -f')   # 시스템 종료

    def _sotpTimer(self):
        self.timer.stop()
        self.tab1Label01.setText('종료')


    def _tab1SliderLabel(self, value):
        self.tab1SliderLabel.setText(str(value))

    # Tab3
    def _InsertTableRow(self):
        maxRowCount = self.tab3Table.rowCount()
        
        ckbox = QCheckBox()
        self.tab3Table.setRowCount(maxRowCount+1)
        self.tab3Table.setCellWidget(maxRowCount, 0, ckbox)


    def _DeleteTableRow(self):
        print()
        # self.tab2Table.removeRow()

    def _print(self):
        # print(self.Line1.text())
        # self.Line2.setText(self.Line1.text())

        for i in range(9):
            # configValueInput["tab1LineEdit"+str(i)] = tab6Entries[i].get()
            print(self.tab2LineEdit[i].text())

    # Tab8
    def _tab8ExcelCreate(self):

        filePath = self.tab8FilePath.text()

        if filePath == "":
            QMessageBox.about(self, '선택', '파일을 선택해 주세요!')
            return
        

        wb = load_workbook(filePath)
        ws = wb['컬럼']
        getData = ws['A2:B13']  # 데이터 범위

        beforeTableName = ''
        nowTableName = ''
        maxDataCnt = len(getData)
        i = 0

        for trow in range(len(getData)):
            nowTableName = ws.cell(trow+2, 1).value
            outFileName = beforeTableName+'.sql'
            
            if  beforeTableName == '' or beforeTableName != nowTableName:
                if beforeTableName != '':
                    fullTextSql = self._tab8FullTextSql(sqlText, beforeTableName)
                    self._tab8CreateFile(fullTextSql, outFileName)
                    i = 0

                sqlColumnText = ''
                sqlText = ''


            sqlColumnText = ws.cell(trow+2, 2).value
            if i == 0:
                sqlText = sqlText + '      ' + sqlColumnText + chr(10)
            else:
                sqlText = sqlText + '    , ' + sqlColumnText + chr(10)

            if maxDataCnt == trow+1:
                fullTextSql = self._tab8FullTextSql(sqlText, beforeTableName)
                self._tab8CreateFile(fullTextSql, outFileName)

            i = i + 1  # 첫번째 컬럼에 (,)를 넣지 않기 위해 체크

            beforeTableName = nowTableName
        # print(sqlFullText)
    def _tab8FullTextSql(self, sqlText, tableName):
        preText = self.tab8TextEdit01.toPlainText()
        postText = self.tab8TextEdit02.toPlainText()

        return preText + chr(10) + chr(10) + 'select ' + chr(10) + sqlText + 'from  ' + tableName + ';' + chr(10) + chr(10) + postText

    def _tab8CreateFile(self, text, fileName):
        fw = open(fileName, 'w')
        fw.write(text)
        fw.close()

    # Tab9
    def _tab9KeywordSearch(self):
        filePath = self.tab9FilePath.text()
        outPathName = os.path.dirname(filePath) + '/' + Path(filePath).stem + '(copy).txt'


        fr = open(filePath, 'r')
        fw = open(outPathName, 'w')

        keywords = self.tab9Keyword.text()

        lines = fr.readlines()
        lines = list(map(lambda s : s.strip(), lines))  # 개행 문자 제거

        results = []

        for i in range(len(lines)):
            if keywords in lines[i] :
                fw.write(lines[i]+'\n')

        fr.close()
        fw.close()


# 자기 자신의 모듈에서 실행할때는 '__main__' 으로 넘어 온다.
# 만약 'moduleA.py'라는 코드를 import해서 예제 코드를 수행하면 __name__ 은 'moduleA'가 된다.
if __name__ == '__main__':     
    app = QApplication(sys.argv)
    gui = MyApp()
    gui.show()
    sys.exit(app.exec_())
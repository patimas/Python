import tkinter as tk
from tkinter import END, Image, StringVar, ttk, Menu, font, IntVar, messagebox

# excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side

#calendar
#pip install tkcalendar
# from tkcalendar import DateEntry


# 파일 저장
import os
import pickle

#DB 연결
# import cx_Oracle

# 이미지, 마우스, 키보드 제어
# import pyautogui as pygui

# ClipBoard
# import pyperclip

############################################################
#
#  GUI 기본값 설정
#
############################################################
win = tk.Tk()
win.title("Python 기본")

# 가로x세로+x축+y축
win.geometry("1160x600+300+300")
# win.resizable(False,False)

###################################################
#                Style
###################################################
style = ttk.Style()

# fNotebook = font.Font(family="고딕체", size="25")
# style.configure("TNotebook", font=fNotebook)

fLabel=font.Font(family="고딕체", size="12")
style.configure("TLabel", font=fLabel)


############################################################
#
#  설정 파일 오픈 (config.ini)
#
############################################################
configPath = "./conf"
configFile = "config.ini"
configFileFullPath = configPath+"/"+configFile

if os.path.isdir(configPath) == False:
    os.mkdir("conf")


if os.path.isfile(configFileFullPath) == True:
    f = open(configFileFullPath,"rb")
    configValueRead = pickle.load(f)
    f.close()



############################################################
#
#  DB접속 쿼리
#
############################################################
# sqlStr0 = "SELECT DISTINCT OWNER,OWNER FROM DBA_TABLES WHERE  OWNER IN ('HR','BTLREP')"

# sqlStr1 = ""
# sqlStr1 +="SELECT  T1.OWNER, T1.TABLE_NAME, T2.COMMENTS      "
# sqlStr1 +="FROM    DBA_TABLES          T1  ,                 "
# sqlStr1 +="        DBA_TAB_COMMENTS    T2                    "
# sqlStr1 +="WHERE   1=1                                       "
# sqlStr1 +="AND     T1.OWNER        = :1                      "
# sqlStr1 +="AND     T1.OWNER        = T2.OWNER                "
# sqlStr1 +="AND     T1.TABLE_NAME   = T2.TABLE_NAME           "
# # sqlStr1 +="AND     T1.TABLE_NAME   LIKE :2                   "
# sqlStr1 +="ORDER BY T1.OWNER, T1.TABLE_NAME        			 "


# sqlStr2 = ""
# sqlStr2 += "SELECT  T1.TABLE_NAME           ,          "
# sqlStr2 += "        T1.COLUMN_NAME          ,          "
# sqlStr2 += "        T2.COMMENTS                        "
# sqlStr2 += "FROM    DBA_TAB_COLUMNS     T1  ,          "
# sqlStr2 += "        DBA_COL_COMMENTS    T2             "
# sqlStr2 += "WHERE   T1.OWNER        = T2.OWNER      (+)"
# sqlStr2 += "AND     T1.TABLE_NAME   = T2.TABLE_NAME (+)"
# sqlStr2 += "AND     T1.COLUMN_NAME  = T2.COLUMN_NAME(+)"
# sqlStr2 += "AND     T1.OWNER        = :1               "
# sqlStr2 += "AND     T1.TABLE_NAME   = :2               "
# sqlStr2 += "ORDER BY T1.COLUMN_ID                      "


############################################################
#
#  데이터 조회
#
############################################################
# def dbConn():
#     global connection
#     global cur
    
#     dsn = cx_Oracle.makedsn("192.168.31.231", 30101, service_name="SADARI")
#     cx_Oracle.init_oracle_client(lib_dir=r"C:\Projects\Python\instantclient_19_9")
#     connection = cx_Oracle.connect(user=configValueRead['tab1_en_user'], password=configValueRead['tab1_en_pass'], dsn=dsn, encoding="UTF-8")
#     # connection = cx_Oracle.connect(user='system', password='oracle', dsn=dsn, encoding="UTF-8")
    
#     cur = connection.cursor()

#     # 화면 생성 시작
#     tab1LFrame01 = ttk.LabelFrame(tab1, text="Table", width=400)
#     tab1LFrame01.grid(row=1, column=0, padx=8, pady=4, sticky=tk.W)

#     # 콤보박스
#     comVar = tk.StringVar()

#     value = on_get_data("combo", sqlStr0, "", "")
#     combo = ttk.Combobox(tab1LFrame01, width=10,  textvariable=comVar, values=value, state='readonly')
#     combo.grid(row=1, column=0)
#     # combo.current(0)

#     # 검색바
#     varStr = StringVar()
#     tab1en1 = ttk.Entry(tab1LFrame01, textvariable=varStr)
#     tab1en1.grid(row=1, column=1)



#     # GIRD 시작
#     tab1LFrame02 = ttk.LabelFrame(tab1, text="Table", width=400)
#     tab1LFrame02.grid(row=2, column=0, padx=8, pady=4)

#     # 1Tree 데이터
#     tab1tv1=ttk.Treeview(tab1LFrame02, columns=["1","2","3"], displaycolumns=["1","2","3"], height=30)
#     tab1tv1.grid(row=3, column=0, columnspan=2)

#     tab1tv1.column("#0", width=70)
#     tab1tv1.heading("#0", text="NO")

#     tab1tv1.column("1",width=100)
#     tab1tv1.heading("1", text="Owner")

#     tab1tv1.column("2",width=200)
#     tab1tv1.heading("2", text="Table")

#     tab1tv1.column("3",width=300)
#     tab1tv1.heading("3", text="Comment")

#     combo.bind('<<ComboboxSelected>>', lambda x : _onComboSelected(combo, tab1tv1, x))
#     # bind가 되었을때 기본적으로 실행 한다.
#     # combo.event_generate('<<ComboboxSelected>>')

#     # 2Tree 데이터
#     tab1tv2 = ttk.Treeview(tab1LFrame02, columns=["1","2","3"], displaycolumns=["1","2","3"], height=30)
#     tab1tv2.grid(row=3, column=2, padx=20, pady=0)

#     tab1tv2.column("#0", width=70)
#     tab1tv2.heading("#0", text="NO")
#     tab1tv2.column("1",width=150)
#     tab1tv2.heading("1", text="Table")
#     tab1tv2.column("2",width=200)
#     tab1tv2.heading("2", text="Column")
#     tab1tv2.column("3",width=300)
#     tab1tv2.heading("3", text="Comment")

#     tab1tv1.bind( '<ButtonRelease-1>', lambda x : on_tab1tv1_click(tab1tv1, tab1tv2 ,x))
#     # tab1tv2.bind( '<Control-Key-c>', lambda x : _copy_clip_board(tab1tv2,x))

# def on_tab1tv1_click(object1, object2, e):
#     selectItem = object1.focus()
#     # getValue = object1.item(selectItem).get('values')

#     getValue = []
#     getValue.append(object1.item(selectItem).get('values')[0])
#     getValue.append(object1.item(selectItem).get('values')[1])

#     on_get_data("sql", sqlStr2, getValue, object2)
    
#     testLabel.configure(text=getValue)


# def on_get_data(flag, sql, param, object):
#     vdata = []

#     if flag == "sql" and param == "":

#         for i in cur.execute(sql):
#             vdata.append(i)

#         for i in range(len(vdata)):
#             object.insert('', 'end', text=i, values=vdata[i], iid=str(i))

#     elif flag == "sql" and param != "":
        
#         # Grid 초기화
#         for i in object.get_children():
#             object.delete(i)

#         for i in cur.execute(sql,param):
#             vdata.append(i)

#         for i in range(len(vdata)):
#             object.insert('', 'end', text=i, values=vdata[i], iid=str(i))

#     elif flag == "combo":
#         cur.execute(sqlStr0)
#         result = [row1[0] for row1 in cur]
#         return result


# def _onComboSelected(object1, object2, e):
#     value = []
#     value.append(object1.get())
#     on_get_data("sql", sqlStr1, value, object2)
#     # on_get_data("sql", sqlStr1, ['HR'], tab1tv1)

# def _copy_clip_board(tree, event):
#     selection = tree.selection()
    
#     column = tree.identify_column(event.x)
    
#     column_no = int(column.replace("#", "")) - 1  #마우스로 선택한 컬럼 index만 가져 오기
#     copy_values = []
#     for each in selection:
#         try:
#             value = tree.item(each)["values"][column_no]
#             copy_values.append(str(value))
#         except:
#             pass
        
#     copy_string = "\n".join(copy_values)
#     pyperclip.copy(copy_string)



############################################################
#
#  이벤트
#
############################################################

# 화면 종료
def _exit():
    # cur.close()
    # connection.close()
    win.destroy()


def _configSave():
    configValueInput = {}

    #Tab Query (1)
    configValueInput['tab1_en_user'] = tab1_en_user.get()
    configValueInput['tab1_en_pass'] = tab1_en_pass.get()

    #Tab Query (7)
    configValueInput['tab7_tx_area1'] = tab7_tx_area1.get(1.0, END)
    configValueInput['tab7_tx_area2'] = tab7_tx_area2.get(1.0, END)
    configValueInput['tab7_tx_area3'] = tab7_tx_area3.get(1.0, END)
    configValueInput['tab7_tx_area4'] = tab7_tx_area4.get(1.0, END)
    configValueInput['tab7_tx_area5'] = tab7_tx_area5.get(1.0, END)
    configValueInput['tab7_tx_area6'] = tab7_tx_area6.get(1.0, END)
    configValueInput['tab7_tx_area7'] = tab7_tx_area7.get(1.0, END)
    configValueInput['tab7_tx_area8'] = tab7_tx_area8.get(1.0, END)
    configValueInput['tab7_tx_area9'] = tab7_tx_area9.get(1.0, END)
    configValueInput['tab7_tx_area10'] = tab7_tx_area10.get(1.0, END)

    # Tab Excel (8)
    configValueInput['tab8Path1'] = tab8_En_Path1.get()
    configValueInput['tab8Path2'] = tab8_En_Path2.get()
    configValueInput['tab8Path3'] = tab8_En_Path3.get()
    # configValueInput['tab8Path4'] = tab8_En_Path4.get()
    configValueInput['tab8Path5'] = tab8_En_Path5.get()
    configValueInput['tab8Path6'] = tab8_En_Path6.get()
    configValueInput['tab8Path7'] = tab8_En_Path7.get()
    configValueInput['tab8Path8'] = tab8_En_Path8.get()
    # configValueInput['tab8Path9'] = tab8_En_Path9.get()
    # configValueInput['tab8Path10'] = tab8_En_Path10.get()


    # Tab Config (9)
    configValueInput['Image_1'] = tab9_En_Path1.get()
    configValueInput['Image_2'] = tab9_En_Path2.get()
    configValueInput['Image_3'] = tab9_En_Path3.get()
    configValueInput['Image_4'] = tab9_En_Path4.get()
    configValueInput['Image_5'] = tab9_En_Path5.get()
    configValueInput['Image_6'] = tab9_En_Path6.get()
    configValueInput['Image_7'] = tab9_En_Path7.get()
    configValueInput['Image_8'] = tab9_En_Path8.get()
    configValueInput['Image_9'] = tab9_En_Path9.get()
    configValueInput['Image_10'] = tab9_En_Path10.get()


    f = open(configFileFullPath,'wb')
    pickle.dump(configFileFullPath, f)
    f.close()

    messagebox.showinfo("저장","저장 완료 되었습니다.!")

def _configSet():

    try:

        #Tab Query (1)
        tab1_en_user.insert(END, configValueRead['tab1_en_user'])
        tab1_en_pass.insert(END, configValueRead['tab1_en_pass'])

        #Tab Query (7)
        tab7_tx_area1.insert(END, configValueRead['tab7_tx_area1'])
        tab7_tx_area2.insert(END, configValueRead['tab7_tx_area2'])
        tab7_tx_area3.insert(END, configValueRead['tab7_tx_area3'])
        tab7_tx_area4.insert(END, configValueRead['tab7_tx_area4'])
        tab7_tx_area5.insert(END, configValueRead['tab7_tx_area5'])
        tab7_tx_area6.insert(END, configValueRead['tab7_tx_area6'])
        tab7_tx_area7.insert(END, configValueRead['tab7_tx_area7'])
        tab7_tx_area8.insert(END, configValueRead['tab7_tx_area8'])
        tab7_tx_area9.insert(END, configValueRead['tab7_tx_area9'])
        tab7_tx_area10.insert(END, configValueRead['tab7_tx_area10'])

        # Tab Config
        tab9_En_Path1.insert(0, configValueRead['Image_1'])
        tab9_En_Path2.insert(0, configValueRead['Image_2'])
        tab9_En_Path3.insert(0, configValueRead['Image_3'])
        tab9_En_Path4.insert(0, configValueRead['Image_4'])
        tab9_En_Path5.insert(0, configValueRead['Image_5'])
        tab9_En_Path6.insert(0, configValueRead['Image_6'])
        tab9_En_Path7.insert(0, configValueRead['Image_7'])
        tab9_En_Path8.insert(0, configValueRead['Image_8'])
        tab9_En_Path9.insert(0, configValueRead['Image_9'])
        tab9_En_Path10.insert(0, configValueRead['Image_10'])

        # Tab Excel (8)
        tab8_En_Path1.insert(0, configValueRead['tab8Path1'])
        tab8_En_Path2.insert(0, configValueRead['tab8Path2'])
        tab8_En_Path3.insert(0, configValueRead['tab8Path3'])
        # tab8_En_Path4.insert(0, configValueRead['tab8Path4'])
        tab8_En_Path5.insert(0, configValueRead['tab8Path5'])
        tab8_En_Path6.insert(0, configValueRead['tab8Path6'])
        tab8_En_Path7.insert(0, configValueRead['tab8Path7'])
        tab8_En_Path8.delete(0,END)                             # Int형일경우 '0'값이 들어가 있기 때무넹 '0'+@ 값으로 나온다;
        tab8_En_Path8.insert(0, configValueRead['tab8Path8'])
        # tab8_En_Path9.insert(0, configValueRead['tab8Path9'])
        # tab8_En_Path10.insert(0, configValueRead['tab8Path10'])



    except:
        print("Error Bye~~~~")
    finally:
        print("End Bye~~~~")


def excelCreate():
    filePath = vartab8Str1.get()
    sourceFile = vartab8Str2.get()
    targetFile = vartab8Str3.get()


    src_wb = openpyxl.load_workbook(filePath+sourceFile)
    tgt_wb = openpyxl.load_workbook(filePath+targetFile)

    src_sheet = src_wb["컬럼"]
    # print(src_sheet['B2'].value)
    # print(src_sheet.cell(3,2).value)

    getData = src_sheet[vartab8Str6.get()]  # 데이터 범위

    sam_sheet = tgt_wb[vartab8Str7.get()] # 샘플 Sheet명

    rowindex = vartab8Str8.get()
    colindex = 1

    beforeTableName = ""
    nowTableName = ""

    i=1

    # Style
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
		                 top=Side(style="thin"), bottom=Side(style="thin"))

    for row in getData:

        i = i+1
        nowTableName = src_sheet.cell(i, 1).value

        if beforeTableName == "" or beforeTableName != nowTableName:
            copy_ws = tgt_wb.copy_worksheet(sam_sheet)
            copy_ws.title = nowTableName
            
            copy_ws["B1"].value = nowTableName

            copy_ws.sheet_view.showGridLines= False  # GridLine 삭제
            rowindex = vartab8Str8.get()
            colindex = 1


        copy_ws.insert_rows(rowindex)

        for cell in row:
            copy_ws.cell(rowindex, colindex).value = cell.value
            copy_ws.cell(rowindex, colindex).border = thin_border
            colindex = colindex +1

        rowindex = rowindex +1
        colindex = 1
        beforeTableName = nowTableName

    tgt_wb.save(filePath+targetFile)
    src_wb.close()
    tgt_wb.close()

    messagebox.showinfo("완료","완료")

def gridHide(val):

    tab7_tx_area1.grid_forget()
    tab7_tx_area2.grid_forget()
    tab7_tx_area3.grid_forget()
    tab7_tx_area4.grid_forget()
    tab7_tx_area5.grid_forget()
    tab7_tx_area6.grid_forget()
    tab7_tx_area7.grid_forget()
    tab7_tx_area8.grid_forget()
    tab7_tx_area9.grid_forget()
    tab7_tx_area10.grid_forget()

    if val == "1":
        tab7_tx_area1.grid(row=0, column=0)
    elif val == "2":
        tab7_tx_area2.grid(row=0, column=0)
    elif val == "3":
        tab7_tx_area3.grid(row=0, column=0)
    elif val == "4":
        tab7_tx_area4.grid(row=0, column=0)
    elif val == "5":
        tab7_tx_area5.grid(row=0, column=0)
    elif val == "6":
        tab7_tx_area6.grid(row=0, column=0)
    elif val == "7":
        tab7_tx_area7.grid(row=0, column=0)
    elif val == "8":
        tab7_tx_area8.grid(row=0, column=0)
    elif val == "9":
        tab7_tx_area9.grid(row=0, column=0)
    elif val == "10":
        tab7_tx_area10.grid(row=0, column=0)

    tab7LFrame01.config(text=val)


# def ImageDoubleClick():
#     screenImage = pygui.screenshot()
#     result = pygui.locate(serFile3, screenImage, confidence=0.9)

#     if result is None :
#         print("없음")
#     else :
#         pygui.doubleClick(result[0], result[1])



############################################################
#
#  Menu 생성
#
############################################################
menu_bar = tk.Menu(win)

file_menu = tk.Menu(menu_bar, tearoff=0)
file_menu.add_command(label="Save", command=_configSave)
file_menu.add_command(label="Exit", command=_exit)
menu_bar.add_cascade(label="FILE", menu=file_menu)  #화면에 그려주는 구문

win.config(menu=menu_bar)

testLabel = ttk.Label(win, text="test")
testLabel.pack()

############################################################
#
#  TAB 생성
#
############################################################
tabCtrl = ttk.Notebook(win)

tab1 = ttk.Frame(tabCtrl)
tab6 = ttk.Frame(tabCtrl)  # 
tab7 = ttk.Frame(tabCtrl)  # sql
tab8 = ttk.Frame(tabCtrl)  # excel
tab9 = ttk.Frame(tabCtrl)  # config

tabCtrl.add(tab1, text="Table")
tabCtrl.add(tab6, text="쭈~~")
tabCtrl.add(tab7, text="Query")
tabCtrl.add(tab8, text="Excel")
tabCtrl.add(tab9, text="Config")
tabCtrl.pack(expand=1, fill="both")


############################################################
#
#  Tab DB (1)
#
############################################################
tab1LFrame00 = ttk.LabelFrame(tab1, text="Connenction", width=400)
tab1LFrame00.grid(row=0, column=0, padx=8, pady=4, sticky=tk.W)
tab1_lb_user = ttk.Label(tab1LFrame00, text="사용자:")
tab1_lb_user.grid(row=0, column=0)

varDbUserStr = StringVar()
tab1_en_user = ttk.Entry(tab1LFrame00, text='', textvariable=varDbUserStr)
tab1_en_user.grid(row=0, column=1, padx=4, pady=2)

tab1_lb_pass = ttk.Label(tab1LFrame00, text="비번:")
tab1_lb_pass.grid(row=0, column=2, padx=4, pady=2)

varDbPassStr = StringVar()
tab1_en_pass = ttk.Entry(tab1LFrame00, textvariable=varDbPassStr)
tab1_en_pass.grid(row=0, column=3, padx=4, pady=2)

# tab1_bn_connect = ttk.Button(tab1LFrame00, text="접속", command=dbConn)
# tab1_bn_connect.grid(row=0, column=4)



############################################################
#
#  Tab DB (6)
#
############################################################
# tab6LFrame00 = ttk.LabelFrame(tab6, text="Entry", width=400)
# tab6LFrame00.grid(row=0, column=0, padx=8, pady=4, sticky=tk.W)

# stringVars = []
# entries = []

# colCount = 7
# rowCount = 10

# tab6_Lb_List_01 = ttk.Label(tab6LFrame00, text="1").grid(row=0, column=0, padx=8, pady=4)
# tab6_Lb_List_02 = ttk.Label(tab6LFrame00, text="2").grid(row=0, column=1, padx=8, pady=4)
# tab6_Lb_List_03 = ttk.Label(tab6LFrame00, text="3").grid(row=0, column=2, padx=8, pady=4)
# tab6_Lb_List_04 = ttk.Label(tab6LFrame00, text="4").grid(row=0, column=3, padx=8, pady=4)
# tab6_Lb_List_05 = ttk.Label(tab6LFrame00, text="5").grid(row=0, column=4, padx=8, pady=4)
# tab6_Lb_List_06 = ttk.Label(tab6LFrame00, text="6").grid(row=0, column=5, padx=8, pady=4)
# tab6_Lb_List_07 = ttk.Label(tab6LFrame00, text="7").grid(row=0, column=6, padx=8, pady=4)

# for i in range(rowCount):
#     for j in range(colCount):
#         stringVar = StringVar()
#         stringVars.append(stringVar)

#         en = ttk.Entry(tab6LFrame00, textvariable=stringVar)
#         en.grid(row=i+1, column=j, padx=8, pady=4)




############################################################
#
#  Tab Qeury (7)
#
############################################################
tab7LFrame00 = ttk.LabelFrame(tab7, text="버튼", width=400)
tab7LFrame00.grid(row=0, column=0, padx=8, pady=4, sticky=tk.N)

tab7_Bn_Str1 = ttk.Button(tab7LFrame00, text="1", width=3, command=lambda : gridHide('1'))
tab7_Bn_Str2 = ttk.Button(tab7LFrame00, text="2", width=3, command=lambda : gridHide('2'))
tab7_Bn_Str3 = ttk.Button(tab7LFrame00, text="3", width=3, command=lambda : gridHide('3'))
tab7_Bn_Str4 = ttk.Button(tab7LFrame00, text="4", width=3, command=lambda : gridHide('4'))
tab7_Bn_Str5 = ttk.Button(tab7LFrame00, text="5", width=3, command=lambda : gridHide('5'))
tab7_Bn_Str6 = ttk.Button(tab7LFrame00, text="6", width=3, command=lambda : gridHide('6'))
tab7_Bn_Str7 = ttk.Button(tab7LFrame00, text="7", width=3, command=lambda : gridHide('7'))
tab7_Bn_Str8 = ttk.Button(tab7LFrame00, text="8", width=3, command=lambda : gridHide('8'))
tab7_Bn_Str9 = ttk.Button(tab7LFrame00, text="9", width=3, command=lambda : gridHide('9'))
tab7_Bn_Str10 = ttk.Button(tab7LFrame00, text="10", width=3, command=lambda : gridHide('10'))
tab7_Bn_Str1.grid(row=0, column=1)
tab7_Bn_Str2.grid(row=0, column=2)
tab7_Bn_Str3.grid(row=0, column=3)
tab7_Bn_Str4.grid(row=0, column=4)
tab7_Bn_Str5.grid(row=0, column=5)
tab7_Bn_Str6.grid(row=1, column=1)
tab7_Bn_Str7.grid(row=1, column=2)
tab7_Bn_Str8.grid(row=1, column=3)
tab7_Bn_Str9.grid(row=1, column=4)
tab7_Bn_Str10.grid(row=1, column=5)

tab7LFrame01 = ttk.LabelFrame(tab7, text="버튼", width=400)
tab7LFrame01.grid(row=0, column=1, padx=8, pady=4, sticky=tk.W)

tab7_tx_area1 = tk.Text(tab7LFrame01)
tab7_tx_area2 = tk.Text(tab7LFrame01)
tab7_tx_area3 = tk.Text(tab7LFrame01)
tab7_tx_area4 = tk.Text(tab7LFrame01)
tab7_tx_area5 = tk.Text(tab7LFrame01)
tab7_tx_area6 = tk.Text(tab7LFrame01)
tab7_tx_area7 = tk.Text(tab7LFrame01)
tab7_tx_area8 = tk.Text(tab7LFrame01)
tab7_tx_area9 = tk.Text(tab7LFrame01)
tab7_tx_area10 = tk.Text(tab7LFrame01)


############################################################
#
#  Tab Excel (8)
#
############################################################
tab8LFrame00 = ttk.LabelFrame(tab8, text="경로", width=400)
tab8LFrame00.grid(row=0, column=0, padx=8, pady=4, sticky=tk.W)


tab8_Lb_Path1 = ttk.Label(tab8LFrame00, text="경로 : ")
tab8_Lb_Path1.grid(row=0, column=0, sticky=tk.E)
vartab8Str1 = StringVar()
tab8_En_Path1 = ttk.Entry(tab8LFrame00, text='', textvariable=vartab8Str1)
tab8_En_Path1.grid(row=0, column=1, padx=4, pady=2)


tab8_Lb_Path2 = ttk.Label(tab8LFrame00, text="소스파일 : ")
tab8_Lb_Path2.grid(row=1, column=0, sticky=tk.E)
vartab8Str2 = StringVar()
tab8_En_Path2 = ttk.Entry(tab8LFrame00, text='', textvariable=vartab8Str2)
tab8_En_Path2.grid(row=1, column=1, padx=4, pady=2)

tab8_Lb_Path3 = ttk.Label(tab8LFrame00, text="타겟파일 : ")
tab8_Lb_Path3.grid(row=2, column=0, sticky=tk.E)
vartab8Str3 = StringVar()
tab8_En_Path3 = ttk.Entry(tab8LFrame00, text='', textvariable=vartab8Str3)
tab8_En_Path3.grid(row=2, column=1, padx=4, pady=2)

# tab8_Lb_Path4 = ttk.Label(tab8LFrame00, text="샘플파일 : ")
# tab8_Lb_Path4.grid(row=3, column=0, sticky=tk.E)
# vartab8Str4 = StringVar()
# tab8_En_Path4 = ttk.Entry(tab8LFrame00, text='', textvariable=vartab8Str4)
# tab8_En_Path4.grid(row=3, column=1, padx=4, pady=2)


tab8LFrame01 = ttk.LabelFrame(tab8, text="Set", width=400)
tab8LFrame01.grid(row=0, column=1, padx=8, pady=4, sticky=tk.W)

tab8_Lb_Path5 = ttk.Label(tab8LFrame01, text="시   작 row : ")
tab8_Lb_Path5.grid(row=0, column=0, sticky=tk.E)
vartab8Str5 = StringVar()
tab8_En_Path5 = ttk.Entry(tab8LFrame01, text='', textvariable=vartab8Str5)
tab8_En_Path5.grid(row=0, column=1, padx=4, pady=2)

tab8_Lb_Path6 = ttk.Label(tab8LFrame01, text="데이터 범위 : ")
tab8_Lb_Path6.grid(row=0, column=2, padx=4, pady=2, sticky=tk.E)
vartab8Str6 = StringVar()
tab8_En_Path6 = ttk.Entry(tab8LFrame01, text='', textvariable=vartab8Str6)
tab8_En_Path6.grid(row=0, column=3, padx=4, pady=2)

tab8_Lb_Path7 = ttk.Label(tab8LFrame01, text="샘플Sheet : ")
tab8_Lb_Path7.grid(row=1, column=0, padx=4, pady=2, sticky=tk.E)
vartab8Str7 = StringVar()
tab8_En_Path7 = ttk.Entry(tab8LFrame01, text='', textvariable=vartab8Str7)
tab8_En_Path7.grid(row=1, column=1, padx=4, pady=2)

tab8_Lb_Path8 = ttk.Label(tab8LFrame01, text="append row : ")
tab8_Lb_Path8.grid(row=2, column=0, padx=4, pady=2, sticky=tk.E)
vartab8Str8 = IntVar()
tab8_En_Path8 = ttk.Entry(tab8LFrame01, text='', textvariable=vartab8Str8)
tab8_En_Path8.grid(row=2, column=1, padx=4, pady=2)


tab8LFrame02 = ttk.LabelFrame(tab8, text="Event", width=400)
tab8LFrame02.grid(row=1, column=0, padx=8, pady=4, sticky=tk.W)

tab8_Bn_Path1 = ttk.Button(tab8LFrame02, text="생성", command=excelCreate)
tab8_Bn_Path1.grid(row=0, column=0, padx=4, pady=2, sticky=tk.W)




############################################################
#
#  Tab Config (9)
#
############################################################
# 접속정보
tab9LFrame00 = ttk.LabelFrame(tab9, text="Config", width=400)
tab9LFrame00.grid(row=0, column=0, padx=8, pady=4, sticky=tk.W)


tab9_Lb_Path1 = ttk.Label(tab9LFrame00, text="이미지1 : ")
tab9_Lb_Path1.grid(row=0, column=0)
vartab9Str1 = StringVar()
tab9_En_Path1 = ttk.Entry(tab9LFrame00, text='', textvariable=vartab9Str1)
tab9_En_Path1.grid(row=0, column=1, padx=4, pady=2)

tab9_Lb_Path2 = ttk.Label(tab9LFrame00, text="이미지2 : ")
tab9_Lb_Path2.grid(row=1, column=0)
vartab9Str2 = StringVar()
tab9_En_Path2 = ttk.Entry(tab9LFrame00, text='', textvariable=vartab9Str2)
tab9_En_Path2.grid(row=1, column=1, padx=4, pady=2)

tab9_Lb_Path3 = ttk.Label(tab9LFrame00, text="이미지3 : ")
tab9_Lb_Path3.grid(row=2, column=0)
vartab9Str3 = StringVar()
tab9_En_Path3 = ttk.Entry(tab9LFrame00, text='', textvariable=vartab9Str3)
tab9_En_Path3.grid(row=2, column=1, padx=4, pady=2)

tab9_Lb_Path4 = ttk.Label(tab9LFrame00, text="이미지4 : ")
tab9_Lb_Path4.grid(row=3, column=0)
vartab9Str4 = StringVar()
tab9_En_Path4 = ttk.Entry(tab9LFrame00, text='', textvariable=vartab9Str4)
tab9_En_Path4.grid(row=3, column=1, padx=4, pady=2)

tab9_Lb_Path5 = ttk.Label(tab9LFrame00, text="이미지5 : ")
tab9_Lb_Path5.grid(row=4, column=0)
vartab9Str5 = StringVar()
tab9_En_Path5 = ttk.Entry(tab9LFrame00, text='', textvariable=vartab9Str5)
tab9_En_Path5.grid(row=4, column=1, padx=4, pady=2)

tab9_Lb_Path6 = ttk.Label(tab9LFrame00, text="이미지6 : ")
tab9_Lb_Path6.grid(row=5, column=0)
vartab9Str6 = StringVar()
tab9_En_Path6 = ttk.Entry(tab9LFrame00, text='', textvariable=vartab9Str6)
tab9_En_Path6.grid(row=5, column=1, padx=4, pady=2)

tab9_Lb_Path7 = ttk.Label(tab9LFrame00, text="이미지7 : ")
tab9_Lb_Path7.grid(row=6, column=0)
vartab9Str7 = StringVar()
tab9_En_Path7 = ttk.Entry(tab9LFrame00, text='', textvariable=vartab9Str7)
tab9_En_Path7.grid(row=6, column=1, padx=4, pady=2)

tab9_Lb_Path8 = ttk.Label(tab9LFrame00, text="이미지8 : ")
tab9_Lb_Path8.grid(row=7, column=0)
vartab9Str8 = StringVar()
tab9_En_Path8 = ttk.Entry(tab9LFrame00, text='', textvariable=vartab9Str8)
tab9_En_Path8.grid(row=7, column=1, padx=4, pady=2)

tab9_Lb_Path9 = ttk.Label(tab9LFrame00, text="이미지9 : ")
tab9_Lb_Path9.grid(row=8, column=0)
vartab9Str9 = StringVar()
tab9_En_Path9 = ttk.Entry(tab9LFrame00, text='', textvariable=vartab9Str9)
tab9_En_Path9.grid(row=8, column=1, padx=4, pady=2)

tab9_Lb_Path10 = ttk.Label(tab9LFrame00, text="이미지10 : ")
tab9_Lb_Path10.grid(row=9, column=0, sticky=tk.W)
vartab9Str10 = StringVar()
tab9_En_Path10 = ttk.Entry(tab9LFrame00, text='', textvariable=vartab9Str10)
tab9_En_Path10.grid(row=9, column=1, padx=4, pady=2)


_configSet()

# 종료 버튼 클릭시 System Tray로 이동
win.protocol('WM_DELETE_WINDOW', _exit)

win.mainloop()